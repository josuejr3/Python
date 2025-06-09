# Utlizando openpyxl
from openpyxl.worksheet.worksheet import Worksheet

# Estudantes tem um conjunto de alunos
# Nome, idade e nota

studentes = [
    ["João", 14, 5.5],
    ["Maria", 13, 9.7],
    ["Luiz", 15, 8.8],
    ["Alberto", 16, 10],
]

from pathlib import Path
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet


ROOT_FOLDER = Path(__file__).parent
WORKBOOK_PATH = ROOT_FOLDER / "aula55.xlsx"

# Arquivo do excel (com várias planilhas dentro)
workbook = Workbook()

# As planilhas presentes em um arquivo excel devem ser tipadas
worksheet: Worksheet = workbook.active

# Criando os cabeçalhos
worksheet.cell(1, 1, 'Nome')
worksheet.cell(1, 2, 'Idade')
worksheet.cell(1, 3, 'Nota')

print(WORKBOOK_PATH)
workbook.save(str(WORKBOOK_PATH))



