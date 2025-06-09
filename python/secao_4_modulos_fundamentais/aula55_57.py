# Utlizando openpyxl
from openpyxl.worksheet.worksheet import Worksheet

# Estudantes tem um conjunto de alunos
# Nome, idade e nota

studentes = [
    ["João", 14, 5.5],
    ["Maria", 13, 9.7],
    ["Luiz", 15, 8.8],
    ["Alberto", 16, 10],
]

from pathlib import Path
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet


ROOT_FOLDER = Path(__file__).parent
WORKBOOK_PATH = ROOT_FOLDER / "aula55-57.xlsx"

# Arquivo do excel (com várias planilhas dentro)
workbook = Workbook()

# As planilhas presentes em um arquivo excel devem ser tipadas
worksheet: Worksheet = workbook.active

# Criando os cabeçalhos
worksheet.cell(1, 1, 'Nome')
worksheet.cell(1, 2, 'Idade')
worksheet.cell(1, 3, 'Nota')

# Passando o restante dos dados para a planilha
for i, student_row in enumerate(studentes, start=2):
    for j, student_col in enumerate(student_row, start=1):
        worksheet.cell(i, j, student_col)

# Para criar uma nova planilha
# Obs: apenas a primeira planilha é que é ativada
# O que eu posso fazer para mudar isso é 'definir a prioridade'
workbook.create_sheet('Planilha2', 0)

sheet_name = 'Planilha2'
# Outra forma de ativar uma planilha do arquivo e fazeno o seguinte
# worksheet: Worksheet = workbook[sheet_name]

# Para a remoção
#workbook.remove(workbook[sheet_name])


# Printando as planilhas dentro do arquivo
print(workbook.sheetnames)


print(WORKBOOK_PATH)
workbook.save(str(WORKBOOK_PATH))

